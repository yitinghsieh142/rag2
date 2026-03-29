import json
import os
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

HEADERS = [
    "商品代號",
    "問題（original)",
    "回答(final)",
    "檢索條款(final)",
    "是否 reAct",
    "initial answer",
    "reviseAnswer1",
    "reviseAnswer2",
    "評估結果(final)",
    "coverage_score(final)",
    "C_fact(final)",
    "C_conf(final)",
    "timestamp",
    "latency_sec",
]


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_GRAY = Side(style="thin", color="D9D9D9")
HEADER_BORDER = Border(bottom=THIN_GRAY)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")


def _safe_float(x: Any) -> Optional[float]:
    try:
        if x is None or x == "":
            return None
        return float(x)
    except Exception:
        return None


def _extract_titles(expanded_docs: List[Dict[str, Any]]) -> str:
    titles = []
    seen = set()

    for d in expanded_docs or []:
        meta = d.get("meta", {}) or {}
        title = str(meta.get("TITLE", "") or "").strip()
        chunk_id = meta.get("CHUNK_ID")

        if title:
            label = title
        elif chunk_id is not None:
            label = f"CHUNK_ID={chunk_id}"
        else:
            label = "(無TITLE)"

        if label not in seen:
            seen.add(label)
            titles.append(label)

    return " | ".join(titles)


def _ensure_workbook(path: str) -> None:
    if os.path.exists(path):
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(HEADERS)

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = HEADER_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        "A": 16,
        "B": 36,
        "C": 48,
        "D": 36,
        "E": 12,
        "F": 40,
        "G": 40,
        "H": 40,
        "I": 24,
        "J": 18,
        "K": 16,
        "L": 16,
        "M": 22,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    wb.save(path)


def append_graph_result_to_excel(state: Dict[str, Any], excel_path: str = "rag_results.xlsx") -> str:
    """
    將 main_graph.py 的最終 state 追加成 Excel 一列。

    需要的 state 欄位（若沒有也可容錯）:
    - product_id
    - query_original
    - answer
    - expanded_docs
    - eval
    - answer_history  # 建議新增，依序放 initial / revise1 / revise2 ...
    - round
    """
    _ensure_workbook(excel_path)
    wb = load_workbook(excel_path)
    ws = wb["Results"]

    eval_obj = state.get("eval", {}) or {}
    nli = eval_obj.get("nli", {}) or {}
    answer_history = state.get("answer_history", []) or []

    initial_answer = answer_history[0] if len(answer_history) >= 1 else ""
    revise_1 = answer_history[1] if len(answer_history) >= 2 else ""
    revise_2 = answer_history[2] if len(answer_history) >= 3 else ""

    has_react = "是" if len(answer_history) >= 2 or int(state.get("round", 0) or 0) > 0 else "否"

    row = [
        state.get("product_id", ""),
        state.get("query_original", ""),
        state.get("answer", ""),
        _extract_titles(state.get("expanded_docs", []) or []),
        has_react,
        initial_answer,
        revise_1,
        revise_2,
        json.dumps(eval_obj, ensure_ascii=False),
        _safe_float(eval_obj.get("coverage_score")),
        _safe_float(nli.get("C_fact", nli.get("factuality"))),
        _safe_float(eval_obj.get("C_conf", eval_obj.get("信心分數"))),
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        _safe_float(state.get("latency")),
    ]

    ws.append(row)
    row_idx = ws.max_row

    # 文字欄位自動換行
    for col_idx in [2, 3, 4, 6, 7, 8, 9]:
        ws.cell(row=row_idx, column=col_idx).alignment = WRAP_TOP

    # 數字格式
    for col_idx in [10, 11, 12, 14]:
        ws.cell(row=row_idx, column=col_idx).number_format = "0.0000"
        ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="right", vertical="top")

    # 條件格式感的簡單標色
    coverage = _safe_float(eval_obj.get("coverage_score"))
    cfact = _safe_float(nli.get("C_fact", nli.get("factuality")))
    cconf = _safe_float(eval_obj.get("C_conf", eval_obj.get("信心分數")))

    def _mark_if_low(col_idx: int, value: Optional[float]):
        if value is not None and value < 0.8:
            ws.cell(row=row_idx, column=col_idx).fill = PatternFill("solid", fgColor="FCE4D6")

    _mark_if_low(10, coverage)
    _mark_if_low(11, cfact)
    _mark_if_low(12, cconf)

    wb.save(excel_path)
    return os.path.abspath(excel_path)


if __name__ == "__main__":
    # 簡單測試
    sample_state = {
        "product_id": "樂鍾溢",
        "query_original": "進行美容手術住院可以申請理賠嗎?",
        "answer": "回答：...",
        "expanded_docs": [
            {"text": "...", "meta": {"CHUNK_ID": 19, "TITLE": "第十八條 除外責任（一）"}},
            {"text": "...", "meta": {"CHUNK_ID": None, "TITLE": "payout_items"}},
        ],
        "answer_history": ["初始答案", "修正版答案1"],
        "round": 1,
        "eval": {
            "coverage_score": 1.0,
            "nli": {"C_fact": 0.9997},
            "C_conf": 0.9998,
        },
    }
    path = append_graph_result_to_excel(sample_state, "rag_results.xlsx")
    print("saved:", path)
